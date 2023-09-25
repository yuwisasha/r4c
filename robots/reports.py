from typing import Any

from datetime import datetime, timedelta

from openpyxl import Workbook
from django.db.models import Count
from django.db.models.manager import BaseManager
from django.db.models.query import QuerySet
from django.conf import settings

from .models import Robot


def write_report(
    wb: Workbook,
    robots_models: QuerySet[Robot, dict[str, Any]],
    robots: BaseManager[Robot],
) -> None:
    for robot in robots_models:
        qs = (
            robots.filter(model=robot["model"])
            .values("model", "version")
            .annotate(total=Count("model"))
            .order_by("total")
        )

        ws = wb.create_sheet(robot["model"])
        ws.append(["Модель", "Версия", "Количество за неделю"])

        for model in qs:
            ws.append([model["model"], model["version"], model["total"]])

        total_cell = ws.cell(row=ws.max_row + 1, column=3)
        total_cell.value = f"=SUM(C2:C{ws.max_row-1})"
        total_cell.font = total_cell.font.copy(bold=True)


def create_report() -> str:
    robots_last_week = Robot.objects.filter(
        created__gte=datetime.now() - timedelta(days=7)
    )
    robots_models = robots_last_week.values("model").distinct()

    wb = Workbook()
    write_report(wb, robots_models, robots_last_week)
    filename = f"{settings.MEDIA_ROOT}/reports/report_{datetime.now().isoformat(timespec='seconds')}.xlsx"  # noqa
    wb.save(filename)
    return filename
